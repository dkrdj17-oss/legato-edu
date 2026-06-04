from fastapi import FastAPI, HTTPException, Body
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import time
import json
import os

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------
# 파일 정의 
# ---------------------------
ERROR_FILE = "errors.json"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ERROR_DB_FILE = os.path.join(BASE_DIR, "error_db.json")

# ---------------------------
# 모델
# ---------------------------
class WBSRequest(BaseModel):
    project: str
    include_error_history: bool = True

# ---------------------------
# 에러 로그
# ---------------------------
def load_errors():
    if not os.path.exists(ERROR_FILE):
        return []
    with open(ERROR_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_error(error):
    errors = load_errors()
    if error not in errors:
        errors.append(error)
    with open(ERROR_FILE, "w", encoding="utf-8") as f:
        json.dump(errors, f, ensure_ascii=False, indent=2)

# ---------------------------
# 에러 DB
# ---------------------------
def load_error_db():
    if not os.path.exists(ERROR_DB_FILE):
        return []
    with open(ERROR_DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

@app.post("/add-error")
def add_error(error: dict):
    db = load_error_db()

    if any(e["code"] == error["code"] for e in db):
        return {"message": "이미 존재하는 코드"}

    db.append(error)

    with open(ERROR_DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

    return {"message": "에러 등록 완료"}

# ---------------------------
# 에러 매핑
# ---------------------------
def attach_error_info(task):
    db = load_error_db()
    task_name = task["task"].lower()

    for e in db:
        keywords = e.get("keywords", [])

        if any(k.lower() in task_name for k in keywords):
            task["error_code"] = e["code"]
            task["risk"] = e["error"]
            task["solution"] = e["solution"]
            return

    task["error_code"] = "-"
    task["risk"] = "일반 작업 리스크"
    task["solution"] = "-"

# ---------------------------
# WBS 생성
# ---------------------------
def generate_wbs_data(project: str, include_error_history: bool):
    errors = load_errors()

    base_tasks = [
        {"wbs": "1", "task": "현행 환경 분석 및 자원 산정", "duration": 1, "start_date": "2026-04-22", "end_date": "2026-04-22", "assignee": "변준석"},
        {"wbs": "2", "task": "네트워크 및 스토리지 구성", "duration": 2, "start_date": "2026-04-23", "end_date": "2026-04-24", "assignee": "김기연"},
        {"wbs": "3", "task": "디스크 변환 및 데이터 이관", "duration": 1, "start_date": "2026-04-25", "end_date": "2026-04-25", "assignee": "신은수"},
        {"wbs": "4", "task": "드라이버 및 시스템 최적화", "duration": 1, "start_date": "2026-04-26", "end_date": "2026-04-26", "assignee": "이종민"},
        {"wbs": "5", "task": "서비스 검증 및 안정화", "duration": 1, "start_date": "2026-04-27", "end_date": "2026-04-27", "assignee": "변준석"},
    ]

    if "Windows" not in project:
        base_tasks[3]["task"] = "Linux 커널 및 GRUB 설정"

    # 에러 매핑
    for task in base_tasks:
        attach_error_info(task)

    # 과거 장애 반영
    if include_error_history and errors:
        error_text = " | ".join(errors[:3])
        for task in base_tasks:
            task["risk"] += f" / 과거 장애: {error_text}"

    return base_tasks

# ---------------------------
# API
# ---------------------------
@app.post("/generate-wbs")
def generate_wbs(req: WBSRequest):
    return {"result": generate_wbs_data(req.project, req.include_error_history)}

# ---------------------------
# 엑셀 생성
# ---------------------------
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def create_excel(data, file_name):
    wb = Workbook()
    ws = wb.active

    headers = ["WBS","작업명","시작일","완료일","작업량(h)","담당자","에러코드","리스크","조치방안"]
    ws.append(headers)

    error_fill = PatternFill(start_color="FFCCCC", fill_type="solid")

    row_idx = 2

    for item in data:
        ws.append([
            item["wbs"],
            item["task"],
            item["start_date"],
            item["end_date"],
            item["duration"] * 8,
            item["assignee"],
            item["error_code"],
            item["risk"],
            item["solution"]
        ])

        # 🔥 하이라이트는 여기
        if item["error_code"] != "-":
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = error_fill

        row_idx += 1

    # 🔥 컬럼 너비는 마지막
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_name)
    return file_name


# ---------------------------
# 엑셀 다운로드
# ---------------------------
@app.post("/generate-wbs-excel")
def generate_wbs_excel(req: WBSRequest):
    try:
        data = generate_wbs_data(req.project, req.include_error_history)

        file_path = f"wbs_{int(time.time())}.xlsx"
        create_excel(data, file_path)

        return FileResponse(
            path=file_path,
            filename="WBS_Result.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

def apply_delay(task):
    if task["error_code"] != "-":
        task["duration"] += 1

        